from kivymd.app import MDApp
from kivymd.uix.label import MDLabel
from kivymd.toast import toast
from kivymd.uix.datatables import MDDataTable
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.filemanager import MDFileManager
from kivymd.uix.textfield import MDTextField

from kivy.lang import Builder
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.config import Config
from kivy.metrics import dp
from kivy.garden.matplotlib.backend_kivyagg import FigureCanvasKivyAgg
from kivy.properties import ObjectProperty
from kivy.properties import NumericProperty
from datetime import datetime
from pathlib import Path
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from scipy.signal import find_peaks
import numpy as np
import os
import minimalmodbus
import time
import serial
from serial.tools import list_ports
import openpyxl
import snap7
import numpy as np
import threading
import time



plt.style.use('bmh')

colors = {
    "Red": {
        "A200": "#EE2222",
        "A500": "#EE2222",
        "A700": "#EE2222",
    },
    "Blue": {
        "200": "#196BA5",
        "500": "#196BA5",
        "700": "#196BA5",
    },
    "Light": {
        "StatusBar": "E0E0E0",
        "AppBar": "#202020",
        "Background": "#FFFFFF",
        "CardsDialogs": "#EEEEEE",
        "FlatButtonDown": "#CCCCCC",
    },
    "Dark": {
        "StatusBar": "101010",
        "AppBar": "#E0E0E0",
        "Background": "#111111",
        "CardsDialogs": "#000000",
        "FlatButtonDown": "#333333",
    },
}

field_pos_left = [
    [0.242,0.790],[0.242,0.990], #lokomotif
    [0.262,0.790],[0.262,0.990],
    [0.290,0.790],[0.290,0.990],
    [0.310,0.790],[0.310,0.990],
    [0.338,0.790],[0.338,0.990],
    [0.358,0.790],[0.358,0.990],

    [0.440,0.790],[0.440,0.990], #gerbong 1
    [0.460,0.790],[0.460,0.990],
    [0.534,0.790],[0.534,0.990],
    [0.554,0.790],[0.554,0.990],

    [0.650,0.790],[0.650,0.990], #gerbong 2
    [0.670,0.790],[0.670,0.990],
    [0.744,0.790],[0.744,0.990],
    [0.764,0.790],[0.764,0.990],

    [0.229,0.530],[0.229,0.730], #gerbong 3
    [0.251,0.530],[0.251,0.730],
    [0.324,0.530],[0.324,0.730],
    [0.346,0.530],[0.346,0.730],

    [0.439,0.530],[0.439,0.730], #gerbong 4
    [0.461,0.530],[0.461,0.730],
    [0.534,0.530],[0.534,0.730],
    [0.556,0.530],[0.556,0.730],

    [0.650,0.530],[0.650,0.730], #gerbong 5
    [0.670,0.530],[0.670,0.730],
    [0.744,0.530],[0.744,0.730],
    [0.764,0.530],[0.764,0.730],

    [0.229,0.270],[0.229,0.470], #gerbong 6
    [0.251,0.270],[0.251,0.470],
    [0.324,0.270],[0.324,0.470],
    [0.346,0.270],[0.346,0.470],

    [0.439,0.270],[0.439,0.470], #gerbong 7
    [0.461,0.270],[0.461,0.470],
    [0.534,0.270],[0.534,0.470],
    [0.556,0.270],[0.556,0.470],

    [0.650,0.270],[0.650,0.470], #gerbong 8
    [0.670,0.270],[0.670,0.470],
    [0.744,0.270],[0.744,0.470],
    [0.764,0.270],[0.764,0.470],

    [0.229,0.010],[0.229,0.205], #gerbong 9
    [0.251,0.010],[0.251,0.205],
    [0.324,0.010],[0.324,0.205],
    [0.346,0.010],[0.346,0.205],

    [0.439,0.010],[0.439,0.205], #gerbong 10
    [0.461,0.010],[0.461,0.205],
    [0.534,0.010],[0.534,0.205],
    [0.556,0.010],[0.556,0.206],

    [0.650,0.010],[0.650,0.205], #gerbong 11
    [0.670,0.010],[0.670,0.205],
    [0.744,0.010],[0.744,0.205],
    [0.764,0.010],[0.764,0.205],
]

field_pos_right = [
    [0.750,0.990],[0.750,0.790], #lokomotif
    [0.730,0.990],[0.730,0.790],
    [0.700,0.990],[0.700,0.790],
    [0.680,0.990],[0.680,0.790],
    [0.650,0.990],[0.650,0.790],
    [0.630,0.990],[0.630,0.790],

    [0.552,0.990],[0.552,0.790], #gerbong 1
    [0.530,0.990],[0.530,0.790],
    [0.455,0.990],[0.455,0.790],
    [0.433,0.990],[0.433,0.790],

    [0.340,0.990],[0.340,0.790], #gerbong 2
    [0.318,0.990],[0.318,0.790],
    [0.245,0.990],[0.245,0.790],
    [0.223,0.990],[0.223,0.790],

    [0.762,0.730],[0.762,0.530], #gerbong 3
    [0.740,0.730],[0.740,0.530],
    [0.667,0.730],[0.667,0.530],
    [0.645,0.730],[0.645,0.530],

    [0.552,0.730],[0.552,0.530], #gerbong 4
    [0.530,0.730],[0.530,0.530],
    [0.455,0.730],[0.455,0.530],
    [0.433,0.730],[0.433,0.530],

    [0.340,0.730],[0.340,0.530], #gerbong 5
    [0.318,0.730],[0.318,0.530],
    [0.245,0.730],[0.245,0.530],
    [0.223,0.730],[0.223,0.530],

    [0.762,0.470],[0.762,0.270], #gerbong 6
    [0.740,0.470],[0.740,0.270],
    [0.667,0.470],[0.667,0.270],
    [0.645,0.470],[0.645,0.270],

    [0.552,0.470],[0.552,0.270], #gerbong 7
    [0.530,0.470],[0.530,0.270],
    [0.455,0.470],[0.455,0.270],
    [0.433,0.470],[0.433,0.270],

    [0.340,0.470],[0.340,0.270], #gerbong 8
    [0.318,0.470],[0.318,0.270],
    [0.245,0.470],[0.245,0.270],
    [0.223,0.470],[0.223,0.270],

    [0.762,0.205],[0.762,0.010], #gerbong 9
    [0.740,0.205],[0.740,0.010],
    [0.667,0.205],[0.667,0.010],
    [0.645,0.205],[0.645,0.010],

    [0.552,0.205],[0.552,0.010], #gerbong 10
    [0.530,0.205],[0.530,0.010],
    [0.455,0.205],[0.455,0.010],
    [0.433,0.205],[0.433,0.010],

    [0.340,0.205],[0.340,0.010], #gerbong 11
    [0.318,0.205],[0.318,0.010],
    [0.245,0.205],[0.245,0.010],
    [0.223,0.205],[0.223,0.010],
]

DEBUG = False

TEMP_OFFSET = 2.5412
TEMP_GAIN = 5.0 * 1000.0 #channge from A to mA with gain

BEARING_TEMP_DATA = 100
NUMBER_OF_BEARINGS = 100

BEARING_TEMP_MIN = 60.5

# Define constants for PLC connection and database read
PLC_IP = '192.168.0.2'
RACK = 0
SLOT = 1
DB_NUMBER = 3
OFFSET1 = 8
OFFSET2 = 404
BYTES_TO_READ1 = 1278
BYTES_TO_READ2 = 1278
SLEEP_DURATION_DATA = 0.1  # seconds
SLEEP_DURATION_TABLE = 0.5
SLEEP_DURATION_DISPLAY = 1.0
REQUEST_TIME_OUT = 5.0

arr_bearing_temps_left = np.zeros(100)
arr_bearing_temps_right = np.zeros(100)

db_bearing_temps = np.zeros([100, 100])
arr_bearing_temps = np.zeros(100)
arr_calc_bearing_temps = np.zeros(200)
calc_bearing_temps = 0.0
dir_left = False
dir_right = False
prev_dir_left = False
prev_dir_right = False

read_sensor_left = False
read_sensor_right = False

counting_wheel = 0
prev_counting_wheel = 0

flag_autosave_data = False
flag_autosave_graph = False
graph_state = 0

class ScreenSplash(MDBoxLayout):
    screen_manager = ObjectProperty(None)
    screen_standby = ObjectProperty(None)
    app_window = ObjectProperty(None)
    
    def __init__(self, **kwargs):
        super(ScreenSplash, self).__init__(**kwargs)
        Clock.schedule_interval(self.update_progress_bar, 0.01)

    def update_progress_bar(self, *args):
        if (self.ids.progress_bar.value + 1) < 100:
            raw_value = self.ids.progress_bar_label.text.split("[")[-1]
            value = raw_value[:-2]
            value = eval(value.strip())
            new_value = value + 1
            self.ids.progress_bar.value = new_value
            self.ids.progress_bar_label.text = "Loading.. [{:} %]".format(new_value)
        else:
            self.ids.progress_bar.value = 100
            self.ids.progress_bar_label.text = "Loading.. [{:} %]".format(100)
            # time.sleep(0.2)
            self.screen_manager.current = "screen_dashboard"
            return False

class ScreenData(MDBoxLayout):
    screen_manager = ObjectProperty(None)

    def __init__(self, **kwargs):
        super(ScreenData, self).__init__(**kwargs)
        self.file_manager = MDFileManager(exit_manager=self.exit_manager, select_path=self.select_path)
        Clock.schedule_once(self.delayed_init, 2)

    def delayed_init(self, dt):
        self.data_tables = MDDataTable(
            use_pagination=True,
            pagination_menu_pos="auto",
            rows_num=5,
            column_data=[("Bearing ", dp(20))]+[(f"T{i}", dp(10)) for i in range(1,101)]
        )

        self.ids.layout_tables.add_widget(self.data_tables)

        self.fig, self.ax = plt.subplots()
        self.fig.tight_layout()
        
        self.ax.set_xlabel("Data No.", fontsize=10)
        self.ax.set_ylabel("Temp. [C]", fontsize=10)

        self.ids.layout_graph.add_widget(FigureCanvasKivyAgg(self.fig))
        try:
            self.connect_to_plc()
            Clock.schedule_interval(self.read_plc, SLEEP_DURATION_DATA)
            toast("PLC is sucessfully connected")
        except:
            Clock.schedule_interval(self.auto_reconnect, REQUEST_TIME_OUT)
            toast("PLC is disconnected")

    def auto_reconnect(self, dt):
        try:
            self.connect_to_plc()
            Clock.schedule_interval(self.read_plc, SLEEP_DURATION_DATA)
            Clock.unschedule(self.auto_reconnect)
            toast("PLC is sucessfully connected")
        except:
            toast("PLC is Ddisconnected, try reconnecting..")

    def reset_data(self):
        global db_bearing_temps
        global arr_calc_bearing_temps

        numbers = np.arange(1,101)       
        db_bearing_temps = np.zeros([100, 100])
        arr_calc_bearing_temps = np.zeros(100)
        numbered_db = np.round(np.vstack((numbers,db_bearing_temps.T)), 1)

        self.data_tables.row_data = numbered_db.T.tolist()

    def open_data(self):
        global db_bearing_temps
        global arr_bearing_temps
        global arr_bearing_temps_left, arr_bearing_temps_right
        # db_bearing_temps = np.array([arr_bearing_temps_left, arr_bearing_temps_right])
        # db_bearing_temps = np.array([[i.value for i in j] for j in dataframe_opened['B2':'CW101']])
        # arr_bearing_temps = arr_bearing_temps_left
        # self.update_table()
        # self.update_graph()
        self.file_manager.show(os.path.expanduser(os.getcwd() + "\data"))  # output manager to the screen
        self.manager_open = True

    def select_path(self, path: str):
        try:
            self.exit_manager(path)
        except:
            toast("error select file path")

    def exit_manager(self, *args):
        global db_bearing_temps
        global arr_bearing_temps
        try: 
            toast("opening data")
            # dataframe = openpyxl.load_workbook(*args)
            # dataframe_opened = dataframe.active

            # # Iterate the loop to read the cell values
            
            # for row in range(1, dataframe_opened.max_row):
            #     for col in dataframe_opened.iter_cols(1, dataframe_opened.max_column):
            #         print(col[row].value)

            # db_bearing_temps = np.array(dataframe_opened)
            # db_bearing_temps = np.array([[i.value for i in j] for j in dataframe_opened['B2':'CW101']])
            # arr_bearing_temps = db_bearing_temps[0]
            # arr_bearing_temps = arr_bearing_temps[arr_bearing_temps != np.array(None)]
            # db_bearing_temps = db_bearing_temps[db_bearing_temps != np.array(None)]
            # db_bearing_temps = [d for d in db_bearing_temps if not(None in d )]

            # print(dataframe_opened)
            # print(db_bearing_temps)
            # print(arr_bearing_temps)

            data_set = np.loadtxt(*args, delimiter=";", encoding=None, skiprows=1)
            db_bearing_temps = data_set.T

            print(db_bearing_temps)
            print(arr_bearing_temps)

            self.update_table()
            self.update_graph()

            self.manager_open = False
            self.file_manager.close()
                   
        except Exception as e:
            print("An exception occurred:", e)
            toast('error open file')
            self.manager_open = False
            self.file_manager.close()

    def connect_to_plc(self):
        global plc

        plc = snap7.client.Client()
        plc.connect(PLC_IP, RACK, SLOT)
        return plc

    def read_from_db(self, plc, db_number, offset, bytes_to_read):
        # print(f"Reading from DB {db_number}, offset {offset}, bytes to read {bytes_to_read}")
        db_bytearray = plc.db_read(db_number, offset, bytes_to_read)
        var1 = snap7.util.get_real(db_bytearray, 0)
        return var1, db_bytearray

    def read_plc(self, dt):
        global plc
        global arr_bearing_temps_left, arr_bearing_temps_right
        global dir_left, dir_right
        global prev_dir_left, prev_dir_right
        global read_sensor_left, read_sensor_right
        global counting_wheel, prev_counting_wheel
        # while True:
        try:
            DB_bytearray = plc.db_read(3,1754,8)
            dir_left = snap7.util.get_bool(DB_bytearray, 0, 0)
            dir_right = snap7.util.get_bool(DB_bytearray, 0, 1)

            DB_bytearray = plc.db_read(3,1748,8)
            read_sensor_left = snap7.util.get_bool(DB_bytearray, 0, 0)
            read_sensor_right = snap7.util.get_bool(DB_bytearray, 0, 1)
           
            DB_bytearray = plc.db_read(3,1216,8)
            counting_wheel = snap7.util.get_int(DB_bytearray, 0)

            # if((dir_left and read_sensor_left) or (dir_right and read_sensor_right)):
            #     self.update_table()

            print("dir left:", dir_left, "dir right:" ,  dir_right, "sens A:", read_sensor_left, "sens B:" , read_sensor_right)

            if ((dir_right or dir_left) and not prev_dir_right and not prev_dir_left):
                Clock.schedule_interval(self.auto_load, SLEEP_DURATION_TABLE)

            if ((prev_dir_right or prev_dir_left) and not dir_left and not dir_right):
                try:
                    self.auto_save_data()
                    self.reset_data()
                    Clock.unschedule(self.auto_load)

                except Exception as e:
                    print("An exception occurred:", e)
                    toast("error save data")
            # start_address = 0x6000  # Starting address of the boolean data
            # num_bytes = 10  # Number of bytes to read

            # # Read the boolean data


            # # Convert the byte array to boolean values
            # bool_data = [bool(x) for x in dir_bytes]

            # # Print the boolean data
            # print("YANG INI --> ", bool_data)


            var1, db_bytearray1 = self.read_from_db(plc, DB_NUMBER, OFFSET1, BYTES_TO_READ1)

            for i in range(0, 99):
                arr_bearing_temps_left[i] = snap7.util.get_real(db_bytearray1, i * 4)

            var1, db_bytearray2 = self.read_from_db(plc, DB_NUMBER, OFFSET2, BYTES_TO_READ2)
            for i in range(0, 99):
                arr_bearing_temps_right[i] = snap7.util.get_real(db_bytearray2, i * 4)

            prev_dir_right = dir_right
            prev_dir_left = dir_left
            prev_counting_wheel = counting_wheel
            


            # print("left:", dir_left, "\t,right:", dir_right)
            # print(arr_bearing_temps_right)
            

        except RuntimeError as e:
            print(f"Error reading PLC data: {e}")

            # time.sleep(SLEEP_DURATION)
    def finding_bearings(self, counting_wheel):
        global db_bearing_temps
        global arr_bearing_temps
        global calc_bearing_temps
        global arr_calc_bearing_temps  

        peaks, _ = find_peaks(db_bearing_temps[counting_wheel], height = BEARING_TEMP_MIN)
        arr_bearing_temps = db_bearing_temps[counting_wheel][db_bearing_temps[counting_wheel] != np.array(None)]
        if arr_bearing_temps[peaks].size == 0:
            calc_bearing_temps = np.max(arr_bearing_temps)
        else:
            calc_bearing_temps = np.max(arr_bearing_temps[peaks])
            # calc_bearing_temps = arr_bearing_temps[peaks][0])

    def auto_load(self, dt):
        global db_bearing_temps
        global arr_bearing_temps
        global arr_bearing_temps_left, arr_bearing_temps_right
        global counting_wheel
        # db_bearing_temps = np.array([arr_bearing_temps_left, arr_bearing_temps_right])
        # db_bearing_temps = np.array([[i.value for i in j] for j in dataframe_opened['B2':'CW101']])
        # arr_bearing_temps = arr_bearing_temps_left
        self.update_table()
        self.ids.text_bearing_num.text = str(counting_wheel)
        self.update_bearing_num()

    def update_table(self):           
        global db_bearing_temps
        global arr_bearing_temps
        global calc_bearing_temps
        global arr_calc_bearing_temps        
        global arr_bearing_temps_left, arr_bearing_temps_right
        global dir_left, dir_right
        global read_sensor_left, read_sensor_right
        global counting_wheel

        numbers = np.arange(1,101)

        if (dir_left == True):
            arr_bearing_temps = arr_bearing_temps_left

        if (dir_right == True):
            arr_bearing_temps = arr_bearing_temps_right

        if((dir_left and read_sensor_left) or (dir_right and read_sensor_right)):        
            db_bearing_temps[counting_wheel] = arr_bearing_temps

            self.finding_bearings(counting_wheel)

            if (dir_left == True):
                arr_calc_bearing_temps[counting_wheel*2] = calc_bearing_temps

            if (dir_right == True):
                arr_calc_bearing_temps[(counting_wheel*2)+1] = calc_bearing_temps
                    
            numbered_db = np.vstack((numbers,np.round(db_bearing_temps.T, 1)))
            try:
                self.data_tables.row_data = numbered_db.T.tolist()
            
            except Exception as e:
                print("An exception occurred:", e)

    def update_graph(self, bearing_num = 1):           
        global db_bearing_temps
        global arr_bearing_temps
        global calc_bearing_temps
        global arr_calc_bearing_temps

        try:
            self.fig, self.ax = plt.subplots()
            self.fig.tight_layout()

            self.finding_bearings(bearing_num - 1)
                       
            self.ax.set_xlabel("n", fontsize=10)
            self.ax.set_ylabel("Temp. [C]", fontsize=10)
            self.ax.set_ylim(0, 200)
            self.ax.set_xlim(0, arr_bearing_temps.size)
            self.ax.plot(arr_bearing_temps)
            # self.ax.plot(peaks, arr_bearing_temps[peaks], "x")
            self.ax.plot(np.zeros_like(arr_bearing_temps) + BEARING_TEMP_MIN, "--", color="gray")

            # self.ids.label_bearing_temp.text = str(calc_bearing_temps)
            self.ids.label_bearing_temp.text = str(np.round(arr_calc_bearing_temps[counting_wheel],2))
            
            self.ids.layout_graph.clear_widgets()
            self.ids.layout_graph.add_widget(FigureCanvasKivyAgg(self.fig))
        
        except Exception as e:
            print("An exception occurred:", e)
            err_msg = str(e)
            toast(err_msg)
    
    def update_bearing_num(self):
        self.update_graph(int(self.ids.text_bearing_num.text))
        
    def sort_on_num(self, data):
        try:
            return zip(
                *sorted(
                    enumerate(data),
                    key=lambda l: l[0][0]
                )
            )
        except:
            toast("Error sorting data")

    def save_data(self):
        global db_bearing_temps
        global arr_bearing_temps

        try:
            # name_file = "\data\\" + self.ids.input_file_name.text + ".xlsx"
            name_file_now = datetime.now().strftime("\data\%d_%m_%Y_%H_%M_%S.csv")
            cwd = os.getcwd()
            # if self.ids.input_file_name.text == "":
            #     disk = cwd + name_file_now
            # else:
            disk = cwd + name_file_now

            header_text = "Roda 1"
            for i in range(1,100):
                header_text = header_text + ';' + "Roda " + str(i) 
            
            with open(disk,"wb") as f:
                np.savetxt(f, db_bearing_temps.T, fmt="%.2f",delimiter=";",header=header_text)

            print("sucessfully save data")
            toast("sucessfully save data")

        except:
            print("error saving data")
            toast("error saving data")

    def auto_save_data(self):
        global db_bearing_temps
        global arr_bearing_temps

        try:
            # name_file = "\data\\" + self.ids.input_file_name.text + ".xlsx"
            name_file_now = datetime.now().strftime("\data\%d_%m_%Y_%H_%M_%S.csv")
            cwd = os.getcwd()
            # if self.ids.input_file_name.text == "":
            #     disk = cwd + name_file_now
            # else:
            disk = cwd + name_file_now

            header_text = "Roda 1"
            for i in range(2,101):
                header_text = header_text + ';' + "Roda " + str(i) 
            
            with open(disk,"wb") as f:
                np.savetxt(f, db_bearing_temps.T, fmt="%.2f",delimiter=";",header=header_text)

            name_file_now = datetime.now().strftime("\data\%d_%m_%Y_%H_%M_%S.csv")
            cwd = 'C:\\Users\\khout\\OneDrive\\Desktop\\HISTORY_DATA'
            # if self.ids.input_file_name.text == "":
            #     disk = cwd + name_file_now
            # else:
            disk = cwd + name_file_now

            header_text = "Roda 1"
            for i in range(2,101):
                header_text = header_text + ';' + "Roda " + str(i) 
            
            with open(disk,"wb") as f:
                np.savetxt(f, db_bearing_temps.T, fmt="%.2f",delimiter=";",header=header_text)

            print("sucessfully save data")
            toast("sucessfully save data")

        except:
            print("error saving data")
            toast("error saving data")

    def screen_dashboard(self):
        self.screen_manager.current = 'screen_dashboard'

    def screen_data(self):
        self.screen_manager.current = 'screen_data'

    def exec_shutdown(self): 
        toast("Shutting down system")
        os.system("shutdown /s /t 1") #for windows os
        # os.system("shutdown -h now") #for linux os

class ScreenDashboard(MDBoxLayout):
    def __init__(self, **kwargs):
        super(ScreenDashboard, self).__init__(**kwargs)
        Clock.schedule_once(self.delayed_init, 3)
        
    def delayed_init(self, dt):
        self.standby()
        Clock.schedule_interval(self.auto_load, SLEEP_DURATION_DISPLAY)

    def auto_load(self, dt):
        global dir_left, dir_right

        if (dir_left == True):
            self.move_left()
            
        if (dir_right == True):
            self.move_right()

        if (dir_right == False and dir_left == False):
            self.standby()

    def move_left(self):
        global field_pos_left
        global arr_calc_bearing_temps

        self.ids.background_image.source = 'asset/kereta_kiri.jpg'

        try:
            self.ids.layout_text_temps.clear_widgets()
            for i in range(1,101):
                field = MDLabel(id=f'T_{i}', 
                                #text=f'{i}', -> Untuk Menampilkan Posisi Data
                                text= f'{np.round(arr_calc_bearing_temps[i-1],1)}', #-> Untuk Menampilkan data suhu bearing
                                theme_text_color= 'Primary' if (arr_calc_bearing_temps[i-1] <= BEARING_TEMP_MIN) else 'Error' ,
                                font_style= 'Caption',
                                pos_hint= {'center_x': (field_pos_left[i-1][0]),'center_y': (field_pos_left[i-1][1])}
                )
                self.ids.layout_text_temps.add_widget(field)

        except Exception as e:
            print("An exception occurred:", e)
            toast('error open screen')

    def move_right(self):
        global field_pos_right
        global arr_calc_bearing_temps

        self.ids.background_image.source = 'asset/kereta_kanan.jpg'

        try:
            self.ids.layout_text_temps.clear_widgets()
            for i in range(1,101):
                field = MDLabel(id=f'T_{i}', 
                                #text=f'{i}', -> Untuk Menampilkan Posisi Data
                                text=f'{np.round(arr_calc_bearing_temps[i-1],1)}', #-> Untuk Menampilkan data suhu bearing
                                theme_text_color= 'Primary' if (arr_calc_bearing_temps[i-1] <= BEARING_TEMP_MIN) else 'Error' ,
                                font_style= 'Caption',
                                pos_hint= {'center_x': (field_pos_right[i-1][0]),'center_y': (field_pos_right[i-1][1])}
                )
                self.ids.layout_text_temps.add_widget(field)

        except Exception as e:
            print("An exception occurred:", e)
            toast('error open screen')

    def standby(self):
        self.ids.background_image.source = 'asset/kereta.png'
        self.ids.layout_text_temps.clear_widgets()

    def screen_dashboard(self):
        self.screen_manager.current = 'screen_dashboard'

    def screen_data(self):
        self.screen_manager.current = 'screen_data'

    def exec_shutdown(self): 
        toast("Shutting down system")
        os.system("shutdown /s /t 1") #for windows os
        # os.system("shutdown -h now") #for linux os

class BearingTemperatureMonitoringApp(MDApp):
    def build(self):
        self.theme_cls.colors = colors
        self.theme_cls.primary_palette = "Blue"
        self.icon = "asset\logo_kai.png" #for windows os
        Window.fullscreen = 'auto'
        Window.borderless = True
        Window.allow_screensaver = True

        screen = Builder.load_file("main.kv")
        return screen

if __name__ == "__main__":
    BearingTemperatureMonitoringApp().run()
